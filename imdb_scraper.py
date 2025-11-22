import os
import platform
import subprocess
import pandas as pd
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ======================================================
#  CLEAN TABLE OUTPUT FORMATTING
# ======================================================
def print_table_header():
    print(f"{'Rank':<6} {'Title':<45} {'Year':<6} {'Rating':<8} {'Movie URL'}")
    print("-" * 130)

def print_table_row(rank, title, year, rating, url):
    print(f"{rank:<6} {title:<45} {year:<6} {rating:<8} {url}")


# ======================================================
#                  MAIN SCRAPER
# ======================================================
def scrape_imdb_top_n(headless=False):

    # -----------------------------
    # USER INPUT
    # -----------------------------
    try:
        n = int(input("Enter the number of top movies to scrape (1-250): "))
        if n < 1 or n > 250:
            print("Invalid range! Defaulting to 10 movies.")
            n = 10
    except:
        print("Invalid input. Defaulting to 10 movies.")
        n = 10

    # -----------------------------
    # DRIVER SETUP
    # -----------------------------
    chromedriver_autoinstaller.install()

    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    driver = webdriver.Chrome(options=options)
    driver.get("https://www.imdb.com/chart/top/")

    # -----------------------------
    # FETCH MOVIE LIST ITEMS
    # -----------------------------
    movies = WebDriverWait(driver, 20).until(
        EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "ul.ipc-metadata-list li.ipc-metadata-list-summary-item")
        )
    )

    movie_list = []

    print_table_header()

    # -----------------------------
    # SCRAPING LOOP
    # -----------------------------
    for rank, movie in enumerate(movies[:n], start=1):
        try:
            title_elem = movie.find_element(By.CSS_SELECTOR, "h3.ipc-title__text")
            title = title_elem.text

            year = movie.find_element(By.CSS_SELECTOR, "span.cli-title-metadata-item").text
            rating = movie.find_element(By.CSS_SELECTOR, "span.ipc-rating-star--rating").text

            link_elem = movie.find_element(By.CSS_SELECTOR, "a.ipc-title-link-wrapper")
            movie_url = link_elem.get_attribute("href")

            poster_elem = movie.find_element(By.CSS_SELECTOR, "img")
            poster_url = poster_elem.get_attribute("src") or poster_elem.get_attribute("loadlate")

        except:
            title = year = rating = movie_url = poster_url = "N/A"

        movie_list.append({
            "Rank": rank,
            "Title": title,
            "Year": year,
            "IMDb Rating": rating,
            "Movie URL": movie_url,
            "Poster URL": poster_url
        })

        print_table_row(rank, title[:45], year, rating, movie_url)

    driver.quit()

    # ======================================================
    # SAVE TO EXCEL
    # ======================================================
    excel_file = f"IMDb_Top_{n}.xlsx"
    df = pd.DataFrame(movie_list)
    df.to_excel(excel_file, index=False)

    # ======================================================
    # FORMAT EXCEL (ALIGNMENT + AUTOWIDTH + HYPERLINKS)
    # ======================================================
    wb = load_workbook(excel_file)
    ws = wb.active

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto column width + center alignment
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 5

    # Add clickable hyperlink in Movie URL column
    url_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Movie URL":
            url_col = idx
            break

    if url_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=url_col)
            if cell.value and cell.value != "N/A":
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    wb.save(excel_file)

    print(f"\n✅ Scraping Completed! Saved {n} movies to {excel_file}")

    # ======================================================
    # AUTO OPEN EXCEL FILE
    # ======================================================
    try:
        if platform.system() == "Windows":
            os.startfile(excel_file)
        elif platform.system() == "Darwin":
            subprocess.call(["open", excel_file])
        else:
            subprocess.call(["xdg-open", excel_file])
    except:
        print("⚠ Unable to auto-open Excel. Please open the file manually.")


# ======================================================
# RUN PROGRAM
# ======================================================
if __name__ == "__main__":
    scrape_imdb_top_n(headless=False)
